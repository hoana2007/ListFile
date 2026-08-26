# -*- coding: utf-8 -*-
"""
Ứng dụng "Quản lý danh sách Thư mục & File" viết bằng Kivy.
Tương thích đa nền tảng: chạy được trên Windows/macOS/Linux và đóng gói thành APK Android (qua Buildozer).

Tính năng tương đương bản Tkinter:
  - Chọn thư mục gốc.
  - Quét 3 trường hợp: có thư mục con / không có thư mục con / thư mục rỗng.
  - Hiển thị người tạo (chủ sở hữu) và thời gian cập nhật lần cuối.
  - Xuất báo cáo Excel (.xlsx) bằng openpyxl.
  - Nhấn đúp (double-tap) vào một dòng để mở thư mục/tệp.
"""

import os
import sys
import time
import platform
import getpass
from pathlib import Path
from datetime import datetime

from kivy.app import App
from kivy.lang import Builder
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.popup import Popup
from kivy.uix.label import Label
from kivy.uix.button import Button
from kivy.uix.filechooser import FileChooserListView
from kivy.uix.scrollview import ScrollView
from kivy.utils import platform as kivy_platform
from kivy.properties import StringProperty
from kivy.core.window import Window

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ghi chú: trên Android, module "android" do python-for-android cung cấp;
# trên máy tính nó không tồn tại nên ta import có điều kiện.
try:
    if kivy_platform == "android":
        from android.permissions import request_permissions, Permission
    else:
        request_permissions = None
        Permission = None
except Exception:  # pragma: no cover
    request_permissions = None
    Permission = None


# ---------------------------------------------------------------------------
# Hàm tiện ích
# ---------------------------------------------------------------------------
def get_file_owner(file_path):
    """Lấy người tạo / chủ sở hữu file (tương thích POSIX, Windows, Android)."""
    try:
        if hasattr(os, "getuid"):
            # Linux / macOS / Android (POSIX)
            return Path(file_path).owner()
        if platform.system() == "Windows":
            try:
                import win32security
                sd = win32security.GetFileSecurity(
                    str(file_path), win32security.OWNER_SECURITY_INFORMATION
                )
                owner_sid = sd.GetFileOwner()
                name, domain, _ = win32security.LookupAccountSid(None, owner_sid)
                return f"{domain}\\{name}" if domain else name
            except Exception:
                return getpass.getuser()
    except Exception:
        pass
    return "N/A"


def format_mtime(path):
    """Định dạng thời gian sửa đổi: dd/mm/YYYY HH:MM:SS."""
    try:
        ts = path.stat().st_mtime
        return datetime.fromtimestamp(ts).strftime("%d/%m/%Y %H:%M:%S")
    except Exception:
        return "N/A"


KV = r"""
<Hang>:
    size_hint_y: None
    height: dp(30)
    spacing: dp(4)
    canvas.before:
        Color:
            rgba: self.bg
        Rectangle:
            pos: self.pos
            size: self.size
    Label:
        text: root.ten
        size_hint: (0.45, 1)
        font_size: '12sp'
        halign: 'left'
        valign: 'middle'
        text_size: self.size
        color: root.mau_chu
        bold: root.la_thu_muc
    Label:
        text: root.chu_so_huu
        size_hint: (0.25, 1)
        font_size: '11sp'
        halign: 'left'
        valign: 'middle'
        text_size: self.size
        color: root.mau_chu
    Label:
        text: root.cap_nhat
        size_hint: (0.30, 1)
        font_size: '11sp'
        halign: 'left'
        valign: 'middle'
        text_size: self.size
        color: root.mau_chu

<MainScreen>:
    orientation: "vertical"
    padding: dp(8)
    spacing: dp(6)

    BoxLayout:
        size_hint_y: None
        height: dp(46)
        Label:
            text: "Quản lý danh sách Thư mục & File"
            font_size: '18sp'
            bold: True
            color: 0.106, 0.212, 0.365, 1

    BoxLayout:
        size_hint_y: None
        height: dp(44)
        spacing: dp(8)
        Button:
            text: "Chon thu muc goc"
            font_size: '14sp'
            background_color: 0.88, 0.96, 1, 1
            on_press: root.mo_dau_chon_thu_muc()
        Button:
            text: "Xuat Excel"
            font_size: '14sp'
            background_color: 0.91, 0.96, 0.91, 1
            on_press: root.xuat_excel()

    Label:
        id: lbl_path
        size_hint_y: None
        height: dp(28)
        text: "Chua chon thu muc nao"
        font_size: '13sp'
        italic: True
        halign: 'left'
        valign: 'middle'
        text_size: self.size
        color: 0.2, 0.2, 0.2, 1

    BoxLayout:
        size_hint_y: None
        height: dp(32)
        spacing: dp(4)
        canvas.before:
            Color:
                rgba: 0.106, 0.212, 0.365, 1
            Rectangle:
                pos: self.pos
                size: self.size
        Label:
            text: "Ten Thu muc / File"
            bold: True
            color: 1, 1, 1, 1
            font_size: '13sp'
            halign: 'left'
            valign: 'middle'
            text_size: self.size
        Label:
            text: "Nguoi tao"
            bold: True
            color: 1, 1, 1, 1
            font_size: '13sp'
            halign: 'left'
            valign: 'middle'
            text_size: self.size
        Label:
            text: "Cap nhat lan cuoi"
            bold: True
            color: 1, 1, 1, 1
            font_size: '13sp'
            halign: 'left'
            valign: 'middle'
            text_size: self.size

    ScrollView:
        do_scroll_x: False
        BoxLayout:
            id: grid_data
            orientation: "vertical"
            size_hint_y: None
            height: self.minimum_height
            spacing: dp(1)

<FileChooserPopup>:
    size_hint: 0.97, 0.92
    title: "Chon thu muc"
    BoxLayout:
        orientation: "vertical"
        FileChooserListView:
            id: fc
            dirselect: True
            on_selection: root.khi_chon(fc.selection)
        BoxLayout:
            size_hint_y: None
            height: dp(46)
            spacing: dp(8)
            Button:
                text: "Chon thu muc nay"
                background_color: 0.88, 0.96, 1, 1
                on_press: root.xac_nhan(fc.path, fc.selection)
            Button:
                text: "Dong"
                on_press: root.dismiss()
"""


class Hang(BoxLayout):
    """Một dòng dữ liệu (thư mục hoặc tệp). Nhấn đúp để mở."""

    ten = StringProperty("")
    chu_so_huu = StringProperty("")
    cap_nhat = StringProperty("")
    duong_dan = StringProperty("")
    la_thu_muc = False
    bg = (1, 1, 1, 1)
    mau_chu = (0.2, 0.2, 0.2, 1)
    _last_tap = 0.0

    def __init__(self, **kwargs):
        self.la_thu_muc = kwargs.pop("la_thu_muc", False)
        self.duong_dan = kwargs.pop("duong_dan", "")
        bg_color = kwargs.pop("bg", (1, 1, 1, 1))
        self.bg = bg_color
        if self.la_thu_muc:
            self.mau_chu = (0.05, 0.28, 0.63, 1)
        super().__init__(**kwargs)

    def on_touch_down(self, touch):
        if self.collide_point(*touch.pos) and self.duong_dan:
            now = time.time()
            if now - self._last_tap < 0.4:
                App.get_running_app().root.mo_item(self.duong_dan)
                self._last_tap = 0.0
            else:
                self._last_tap = now
        return super().on_touch_down(touch)


class FileChooserPopup(Popup):
    """Hộp thoại chọn thư mục (dùng FileChooserListView ở chế độ dirselect)."""

    def khi_chon(self, selection):
        pass

    def xac_nhan(self, path, selection):
        # Nếu người dùng chọn một thư mục cụ thể thì dùng nó, ngược lại dùng thư mục hiện tại.
        if selection and os.path.isdir(selection[0]):
            chosen = selection[0]
        else:
            chosen = path
        if not chosen:
            return
        self.dismiss()
        App.get_running_app().root.nhan_thu_muc(chosen)


class MainScreen(BoxLayout):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.current_folder_path = ""
        self.scanned_data = []

    # ----------------------- Quyền truy cập (Android) -----------------------
    def yeu_cau_quyen(self):
        """Yêu cầu quyền truy cập bộ nhớ trên Android (best-effort)."""
        if kivy_platform != "android" or request_permissions is None:
            return
        try:
            perms = [Permission.READ_EXTERNAL_STORAGE, Permission.WRITE_EXTERNAL_STORAGE]
            # Android 11+ (API 30+) cần quyền "All files access" để duyệt toàn bộ bộ nhớ.
            if hasattr(Permission, "MANAGE_EXTERNAL_STORAGE"):
                perms.append(Permission.MANAGE_EXTERNAL_STORAGE)
            request_permissions(perms)
        except Exception:
            pass

    # ----------------------- Chọn thư mục -----------------------
    def mo_dau_chon_thu_muc(self):
        self.yeu_cau_quyen()
        popup = FileChooserPopup()
        popup.open()

    def nhan_thu_muc(self, folder_selected):
        self.current_folder_path = folder_selected
        self.ids.lbl_path.text = f"Duong dan: {folder_selected}"
        self.scanned_data = self.quet_thu_muc(folder_selected)
        self.hien_thi(self.scanned_data)

    # ----------------------- Quét thư mục -----------------------
    def quet_thu_muc(self, folder_selected):
        root_path = Path(folder_selected)
        try:
            items = list(root_path.iterdir())
        except Exception as e:
            self.thong_bao("Loi", f"Khong the truy cap thu muc:\n{e}")
            return []

        sub_folders = [i for i in items if i.is_dir()]
        direct_files = [i for i in items if i.is_file()]
        scanned = []

        # TRƯỜNG HỢP 1: Có các thư mục con
        if sub_folders:
            for item in sub_folders:
                folder_record = {
                    "name": item.name,
                    "owner": get_file_owner(item),
                    "mtime": format_mtime(item),
                    "path": str(item),
                    "files": [],
                    "root": False,
                }
                try:
                    sub_files = [f for f in item.iterdir() if f.is_file()]
                    files_sorted = sorted(
                        sub_files, key=lambda f: f.stat().st_mtime, reverse=True
                    )[:10]
                    for f in files_sorted:
                        folder_record["files"].append({
                            "name": f.name,
                            "owner": get_file_owner(f),
                            "mtime": format_mtime(f),
                            "path": str(f),
                        })
                except Exception as e:
                    folder_record["error"] = str(e)
                scanned.append(folder_record)

        # TRƯỜNG HỢP 2: Không có thư mục con -> liệt kê 10 file mới nhất
        elif direct_files:
            root_record = {
                "name": f"[Thu muc goc] {root_path.name}",
                "owner": get_file_owner(root_path),
                "mtime": format_mtime(root_path),
                "path": str(root_path),
                "files": [],
                "root": True,
            }
            files_sorted = sorted(
                direct_files, key=lambda f: f.stat().st_mtime, reverse=True
            )[:10]
            for f in files_sorted:
                root_record["files"].append({
                    "name": f.name,
                    "owner": get_file_owner(f),
                    "mtime": format_mtime(f),
                    "path": str(f),
                })
            scanned.append(root_record)

        # TRƯỜNG HỢP 3: Thư mục trống hoàn toàn
        else:
            self.thong_bao(
                "Thong bao",
                "Thu muc da chon trong (khong chua thu muc con va khong co file nao).",
            )

        return scanned

    # ----------------------- Hiển thị -----------------------
    def hien_thi(self, scanned):
        grid = self.ids.grid_data
        grid.clear_widgets()
        zebra = False
        for group in scanned:
            base_bg = (0.89, 0.95, 1, 1) if group.get("root") else (0.95, 0.98, 1, 1)
            prefix = "" if group.get("root") else "  "
            icon = "" if group.get("root") else "📁 "
            self._them_hang(
                grid,
                f"{icon}{prefix}{group['name']}",
                group["owner"],
                group["mtime"],
                la_thu_muc=True,
                duong_dan=group.get("path", ""),
                bg=base_bg,
            )
            if group.get("error"):
                self._them_hang(grid, f"   ⚠ Loi: {group['error']}", "", "", bg=(1, 0.9, 0.9, 1))
            elif not group["files"]:
                self._them_hang(grid, "   (Khong co file)", "", "", bg=(0.97, 0.97, 0.97, 1))
            else:
                for f in group["files"]:
                    zebra = not zebra
                    bg = (1, 1, 1, 1) if zebra else (0.97, 0.99, 1, 1)
                    self._them_hang(
                        grid,
                        f"   📄 {f['name']}",
                        f["owner"],
                        f["mtime"],
                        la_thu_muc=False,
                        duong_dan=f.get("path", ""),
                        bg=bg,
                    )

    def _them_hang(self, grid, ten, chu_so_huu, cap_nhat, la_thu_muc=False,
                   duong_dan="", bg=(1, 1, 1, 1)):
        row = Hang(
            ten=ten,
            chu_so_huu=chu_so_huu,
            cap_nhat=cap_nhat,
            la_thu_muc=la_thu_muc,
            duong_dan=duong_dan or "",
            bg=bg,
        )
        grid.add_widget(row)

    # ----------------------- Mở thư mục / tệp -----------------------
    def mo_item(self, path):
        if not os.path.exists(path):
            self.thong_bao("Loi", f"Duong dan khong ton tai:\n{path}")
            return
        if kivy_platform == "android":
            self._mo_android(path)
        else:
            self._mo_desktop(path)

    def _mo_desktop(self, path):
        try:
            if platform.system() == "Windows":
                os.startfile(path)
            elif platform.system() == "Darwin":
                import subprocess
                subprocess.run(["open", path], check=False)
            else:
                import subprocess
                subprocess.run(["xdg-open", path], check=False)
        except Exception as e:
            self.thong_bao("Loi", f"Khong the mo:\n{path}\n{e}")

    def _mo_android(self, path):
        # Best-effort: mở tệp bằng ứng dụng mặc định qua Intent.
        try:
            from jnius import autoclass
            PythonActivity = autoclass("org.kivy.android.PythonActivity")
            Intent = autoclass("android.content.Intent")
            Uri = autoclass("android.net.Uri")
            File = autoclass("java.io.File")

            activity = PythonActivity.mActivity
            f = File(path)
            uri = Uri.fromFile(f)
            intent = Intent(Intent.ACTION_VIEW)
            intent.setDataAndType(uri, self._mime_android(path))
            intent.addFlags(0x10000000)  # FLAG_ACTIVITY_NEW_TASK
            activity.startActivity(intent)
        except Exception as e:
            self.thong_bao(
                "Loi",
                f"Khong the mo tren Android (co the bi han che boi Scoped Storage):\n{e}",
            )

    @staticmethod
    def _mime_android(path):
        ext = os.path.splitext(path)[1].lower()
        mapping = {
            ".txt": "text/plain",
            ".pdf": "application/pdf",
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".mp3": "audio/mpeg",
            ".mp4": "video/mp4",
            ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            ".zip": "application/zip",
        }
        return mapping.get(ext, "*/*")

    # ----------------------- Xuất Excel -----------------------
    def xuat_excel(self):
        if not self.scanned_data:
            self.thong_bao("Canh bao", "Vui long chon thu muc va quet du lieu truoc khi xuat Excel!")
            return

        folder = self.current_folder_path
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"bao_cao_thu_muc_{ts}.xlsx"
        file_path = os.path.join(folder, file_name)

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Danh sach File"

            font_title = Font(name="Arial", size=14, bold=True, color="1B365D")
            font_header = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            font_folder = Font(name="Arial", size=11, bold=True, color="0D47A1")
            font_file = Font(name="Arial", size=10, color="333333")

            fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
            fill_folder = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            fill_zebra = PatternFill(start_color="F9FBFD", end_color="F9FBFD", fill_type="solid")

            thin_border = Border(
                left=Side(style="thin", color="E0E0E0"),
                right=Side(style="thin", color="E0E0E0"),
                top=Side(style="thin", color="E0E0E0"),
                bottom=Side(style="thin", color="E0E0E0"),
            )

            ws.merge_cells("A1:D1")
            ws["A1"] = "BAO CAO DANH SACH THU MUC VA FILE MOI NHAT"
            ws["A1"].font = font_title
            ws["A1"].alignment = Alignment(vertical="center")

            ws["A2"] = f"Thu muc goc: {self.current_folder_path}"
            ws["A2"].font = Font(name="Arial", size=10, italic=True, color="555555")

            headers = ["Thu muc con / Ten File", "Loai", "Nguoi tao (Tac gia)", "Thoi gian cap nhat lan cuoi"]
            ws.append([])
            ws.append(headers)

            for col_num in range(1, 5):
                cell = ws.cell(row=4, column=col_num)
                cell.font = font_header
                cell.fill = fill_header
                cell.alignment = Alignment(horizontal="center" if col_num > 1 else "left", vertical="center")

            current_row = 5
            for item_group in self.scanned_data:
                if not item_group["name"].startswith("[Thu muc goc]"):
                    ws.cell(row=current_row, column=1, value=f"📁 {item_group['name']}").font = font_folder
                    ws.cell(row=current_row, column=2, value="Thu muc con").font = font_folder
                    ws.cell(row=current_row, column=3, value=item_group["owner"]).font = font_folder
                    ws.cell(row=current_row, column=4, value=item_group["mtime"]).font = font_folder
                    for col_num in range(1, 5):
                        c = ws.cell(row=current_row, column=col_num)
                        c.fill = fill_folder
                        c.border = thin_border
                        if col_num in (2, 3, 4):
                            c.alignment = Alignment(horizontal="center", vertical="center")
                    current_row += 1

                if not item_group["files"]:
                    ws.cell(row=current_row, column=1, value="   (Khong co file)").font = font_file
                    for col_num in range(1, 5):
                        ws.cell(row=current_row, column=col_num).border = thin_border
                    current_row += 1
                else:
                    for idx, f in enumerate(item_group["files"]):
                        ws.cell(row=current_row, column=1, value=f"📄 {f['name']}").font = font_file
                        ws.cell(row=current_row, column=2, value="Tep tin").font = font_file
                        ws.cell(row=current_row, column=3, value=f["owner"]).font = font_file
                        ws.cell(row=current_row, column=4, value=f["mtime"]).font = font_file
                        for col_num in range(1, 5):
                            c = ws.cell(row=current_row, column=col_num)
                            c.border = thin_border
                            if idx % 2 == 1:
                                c.fill = fill_zebra
                            if col_num in (2, 3, 4):
                                c.alignment = Alignment(horizontal="center", vertical="center")
                        current_row += 1

            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.row < 4:
                        continue
                    val_str = str(cell.value or "")
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

            wb.save(file_path)
            self.thong_bao("Thanh cong", f"Da xuat file Excel tai:\n{file_path}")
        except Exception as e:
            self.thong_bao("Loi", f"Khong the xuat file Excel:\n{e}")

    # ----------------------- Tiện ích giao diện -----------------------
    def thong_bao(self, tieu_de, noi_dung):
        content = BoxLayout(orientation="vertical", spacing=dp_v(8))
        content.add_widget(Label(text=noi_dung, text_size=(Window.width * 0.8, None),
                                 halign="left", valign="top"))
        btn = Button(text="Dong", size_hint_y=None, height=dp_v(40))
        popup = Popup(title=tieu_de, content=content, size_hint=(0.85, 0.5))
        btn.bind(on_press=popup.dismiss)
        content.add_widget(btn)
        popup.open()


def dp_v(x):
    from kivy.metrics import dp
    return dp(x)


class QuanLyApp(App):
    def build(self):
        self.title = "Quan ly Thu muc & File"
        Builder.load_string(KV)
        return MainScreen()


if __name__ == "__main__":
    QuanLyApp().run()
